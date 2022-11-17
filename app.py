from flask import Flask, render_template
from openpyxl import load_workbook
import os, json, requests, warnings, csv
from datetime import date

basedir = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)
app.config['CURRENTROOM'] = 'B112'
warnings.simplefilter(action='ignore', category=UserWarning)
wb = load_workbook(os.path.join(basedir, 'CNAP_utrustning_inventarie.xlsx'))
ws = wb['Inventarie']

def verify_scantoken(token_type, token):
    '''verifies the token or key against the user json file'''
    if len(token) > 0:
        with open(os.path.join(basedir, 'users.crd')) as file:
            json_data = json.load(file)
            for user in json_data['user_tokens']:
                if token_type == 'IFTTT_key':
                    return user[token_type]
                else:
                    if user[token_type] == token:
                        return True
    return False

def ifttt_message(payload, key):
    msg_data = {}
    msg_data['value1'] = payload
    requests.post(f'https://maker.ifttt.com/trigger/notify/with/key/{key}', data=msg_data)

def verify_barcode(payload):
    if len(payload) == 9:
        # check if its a CNAP qr code
        if payload[0:4] == 'CNAP' and payload[4:].isnumeric():
            return True, 'CNAP'
        elif payload[0:2] == 'IT' and payload[2:].isnumeric():
            return True, 'IT'
        else:
            return False, 'invalid'
    else:
        # neither of these, or corrupt data
        return False, 'invalid'

def cnap_to_itnumber(cnapdata):
    for row in ws['D']:
        if cnapdata == row.value:
            it_cell = ws[f'E{row.row}'].value
            if it_cell is None:
                return 'Not found'
            if len(it_cell) > 0:
                if 'IT' in it_cell:
                    return it_cell
    return 'Not found'

def today():
    today = date.today()
    return today.strftime('%Y-%m-%d')

def add_resource_to_csv(itnr):
    add_header = False if os.path.exists(f'{app.config["CURRENTROOM"]}.csv') else True
    with open(f'{app.config["CURRENTROOM"]}.csv', 'a', newline='') as file:
        writer = csv.writer(file)
        data = [today(), itnr.rstrip('\n')]
        if add_header:
            writer.writerow(['Datum', 'Resurs'])
        writer.writerow(data)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/scan/<token>/<data>')
def scan(token='', data=''):
    if verify_scantoken('scan_token', token):
        # if data is a tuple with a room/number: ie. ('set_room', 'b112')
        if 'set_room' in data:
            from flask import current_app
            room = current_app.config['CURRENTROOM'] = data.split(':')[1]
            ifttt_message(f'room changed to {room}', verify_scantoken('IFTTT_key', 'NULL'))
            return f'room changed to {room}'
        elif 'IT' in data or 'CNAP' in data:
            # CNAP and ITnr, check it and process if valid
            result, verify = verify_barcode(data)
            if result:
                if verify == 'CNAP':
                    itnr = cnap_to_itnumber(data)
                    if len(itnr) > 0:
                        add_resource_to_csv(itnr)
                        return itnr
                    else:
                        return 'No mapping of CNAP-to-IT'
                else:
                    add_resource_to_csv(data)
                    return data
            else:
                return f'{data} invalid'
        else:
            return f'{data} invalid'
    else:
        return 'invalid token'

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')